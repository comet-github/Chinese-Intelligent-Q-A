import openpyxl
from fenci1 import fenci


# 逐行读取每一列的单元格内容中的数值或文本，并将每一行作为一组数据存储
def excelGetData(excelName, sheetName):
    # 打开 Excel 文件
    wb = openpyxl.load_workbook(excelName)
    # 选择要读取的工作表
    sheet = wb[sheetName]

    # 获取最大行数和列数
    max_row = sheet.max_row
    max_column = sheet.max_column
    data = []
    for row_num in range(2, max_row + 1):
        row_data = []
        for col_num in range(1, max_column + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            cell_value = cell.value  # 获取单元格的值（数值或文本而非公式）
            row_data.append(cell_value)
        data.append(row_data)
    return data


# 分词、删除单个汉字、汇总
def row_splitSum(str_list, fenciMethod):
    wards_list = []
    ordnum = -1
    for string in str_list:
        if string is not None and isinstance(string, str):
            words = fenciMethod(string)
            # print("words", words)
            wards_list.extend(words)
            # print("words_set", words_set)
        elif isinstance(string, int) and ordnum == -1:
            ordnum = string

    output_list = []
    if ordnum == -1:
        print("数据ordnum属性值缺失，请补充后重试")
        return output_list
    else:
        output_list.append([ordnum])
        output_list.append(wards_list)
        return output_list


def fenci_allExcelData(excelName, sheetName, fenciMethod):
    data = excelGetData(excelName, sheetName)
    dataList = []
    for d in data:
        # print(d)
        output_list = row_splitSum(d, fenciMethod)
        # print(output_list)
        dataList.append(output_list)
    return dataList


if __name__ == '__main__':
    data = fenci_allExcelData('List of FAQ 20220711.xlsx', 'Sheet2', fenci)
    print(data[0])
